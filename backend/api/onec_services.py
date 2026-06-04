import hashlib
import logging
import re
from collections import Counter, defaultdict
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.db.models import Count
from django.utils import timezone
from openpyxl import load_workbook

from .models import OneCEvent, OneCImportBatch

logger = logging.getLogger(__name__)

VLADIVOSTOK_TZ = timezone.get_fixed_timezone(600)
INVOICE_NUMBER_RE = re.compile(r'([A-ZА-ЯЁ]{1,2}\d?-\d{5}/\d{2})', re.IGNORECASE)
OWN_LAB_PREFIXES = {'В3', 'В4', 'В6'}

CONFUSABLE_LATIN_TO_CYRILLIC = str.maketrans({
    'A': 'А', 'B': 'В', 'C': 'С', 'E': 'Е', 'H': 'Н', 'K': 'К', 'M': 'М',
    'O': 'О', 'P': 'Р', 'T': 'Т', 'X': 'Х', 'Y': 'У',
})

BUSINESS_STATUS_CODES = [
    OneCEvent.StatusCode.ACCEPTED_WAREHOUSE,
    OneCEvent.StatusCode.TRANSFERRED_TO_LAB,
    OneCEvent.StatusCode.ACCEPTED_LAB,
    OneCEvent.StatusCode.TRANSFERRED_FROM_LAB,
    OneCEvent.StatusCode.DONE_READY,
    OneCEvent.StatusCode.ISSUED_CUSTOMER,
]

STATUS_RULES = [
    (OneCEvent.StatusCode.ACCEPTED_WAREHOUSE, ('принято на склад',)),
    (OneCEvent.StatusCode.TRANSFERRED_TO_LAB, ('передано в лабораторию',)),
    (OneCEvent.StatusCode.TRANSFERRED_FROM_LAB, ('передано из лаборатории',)),
    (OneCEvent.StatusCode.ACCEPTED_LAB, ('принято в лаборатории', 'принято в лабораторию')),
    (OneCEvent.StatusCode.DONE_READY, ('выполнено',)),
    (OneCEvent.StatusCode.ISSUED_CUSTOMER, ('выдан заказчику', 'выдано заказчику')),
]

DATE_FORMATS = (
    '%d %m %y %H:%M',
    '%d %m %Y %H:%M',
    '%d.%m.%y %H:%M',
    '%d.%m.%Y %H:%M',
    '%d.%m.%Y %H:%M:%S',
    '%Y-%m-%d %H:%M:%S',
    '%Y-%m-%d %H:%M',
)

PERIOD_LABELS = {
    'last_7': 'последние 7 дней',
    'last_30': 'последние 30 дней',
    'current_week': 'текущая неделя',
    'current_quarter': 'текущий квартал',
    'current_year': 'текущий год',
}

WEEKDAY_SHORT = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']


def normalize_text(value) -> str:
    if value is None:
        return ''
    return str(value).strip()


def normalize_for_hash(value) -> str:
    return ' '.join(normalize_text(value).upper().split())


def normalize_invoice_text(value) -> str:
    text = normalize_text(value).upper().translate(CONFUSABLE_LATIN_TO_CYRILLIC)
    return ' '.join(text.split())


def extract_invoice_number(raw_value: str):
    text = normalize_invoice_text(raw_value)
    match = INVOICE_NUMBER_RE.search(text)
    if not match:
        return '', '', False
    number = match.group(1).upper()
    prefix = number.split('-', 1)[0]
    return number, prefix, prefix in OWN_LAB_PREFIXES


def classify_status(raw_status: str) -> str:
    normalized = normalize_text(raw_status).lower()
    for status_code, needles in STATUS_RULES:
        if any(needle in normalized for needle in needles):
            return status_code
    return OneCEvent.StatusCode.OTHER


def parse_event_datetime(value):
    if value is None or normalize_text(value) == '':
        return None
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, time.min)
    else:
        text = normalize_text(value).replace(',', ' ')
        text = re.sub(r'\s+', ' ', text)
        parsed = None
        for fmt in DATE_FORMATS:
            try:
                parsed = datetime.strptime(text, fmt)
                break
            except ValueError:
                continue
        if parsed is None:
            return None
    if timezone.is_naive(parsed):
        return timezone.make_aware(parsed, VLADIVOSTOK_TZ)
    return parsed.astimezone(VLADIVOSTOK_TZ)


def parse_quantity(value) -> Decimal:
    text = normalize_text(value).replace(',', '.')
    if not text:
        return Decimal('1')
    try:
        quantity = Decimal(text)
    except InvalidOperation:
        return Decimal('1')
    if quantity <= 0:
        return Decimal('1')
    return quantity


def build_instrument_key(invoice_number, barcode, service_name, work_description):
    barcode = normalize_for_hash(barcode)
    description_hash = hashlib.sha1(
        f'{normalize_for_hash(service_name)}|{normalize_for_hash(work_description)}'.encode('utf-8')
    ).hexdigest()[:16]
    if barcode:
        return f'{invoice_number}|barcode:{barcode}|{description_hash}'
    return f'{invoice_number}|desc:{description_hash}'


def make_dedupe_hash(payload: dict) -> str:
    parts = [
        payload.get('event_datetime').isoformat() if payload.get('event_datetime') else '',
        payload.get('invoice_raw', ''),
        payload.get('invoice_number', ''),
        payload.get('counterparty', ''),
        payload.get('consignee', ''),
        payload.get('service_name', ''),
        payload.get('work_description', ''),
        str(payload.get('quantity', '')),
        payload.get('status_raw', ''),
        payload.get('barcode', ''),
        payload.get('responsible_full_name', ''),
    ]
    normalized = '\u241f'.join(normalize_for_hash(part) for part in parts)
    return hashlib.sha256(normalized.encode('utf-8')).hexdigest()


def is_header_row(values) -> bool:
    if not values:
        return False
    first = normalize_text(values[0]).lower()
    second = normalize_text(values[1]).lower() if len(values) > 1 else ''
    return first == 'дата' and ('сч' in second or 'номер' in second)


def parse_row(values, row_number):
    values = list(values or [])
    if not any(normalize_text(value) for value in values):
        return None, None

    if is_header_row(values):
        return None, None

    # Основной формат: 10 колонок с грузополучателем.
    # Примерный старый/короткий формат: 9 колонок без грузополучателя.
    if len(values) >= 10:
        raw_date, invoice_raw, counterparty, consignee, service_name, work_description, quantity, status_raw, barcode, responsible = values[:10]
    elif len(values) >= 9:
        raw_date, invoice_raw, counterparty, service_name, work_description, quantity, status_raw, barcode, responsible = values[:9]
        consignee = ''
    else:
        return None, f'Строка {row_number}: ожидалось 9 или 10 колонок, получено {len(values)}.'

    event_datetime = parse_event_datetime(raw_date)
    if not event_datetime:
        return None, f'Строка {row_number}: не удалось разобрать дату "{raw_date}".'

    invoice_number, invoice_prefix, is_own_lab = extract_invoice_number(invoice_raw)
    if not invoice_number:
        return None, f'Строка {row_number}: не найден номер счёта в поле "{invoice_raw}".'

    quantity_value = parse_quantity(quantity)
    payload = {
        'event_datetime': event_datetime,
        'invoice_raw': normalize_text(invoice_raw),
        'invoice_number': invoice_number,
        'invoice_prefix': invoice_prefix,
        'is_own_lab': is_own_lab,
        'counterparty': normalize_text(counterparty),
        'consignee': normalize_text(consignee),
        'service_name': normalize_text(service_name),
        'work_description': normalize_text(work_description),
        'quantity': quantity_value,
        'status_raw': normalize_text(status_raw),
        'status_code': classify_status(status_raw),
        'barcode': normalize_text(barcode),
        'responsible_full_name': normalize_text(responsible),
        'source_row_number': row_number,
    }
    payload['instrument_key'] = build_instrument_key(
        invoice_number=payload['invoice_number'],
        barcode=payload['barcode'],
        service_name=payload['service_name'],
        work_description=payload['work_description'],
    )
    payload['dedupe_hash'] = make_dedupe_hash(payload)
    return payload, None


@transaction.atomic
def import_onec_xlsx(uploaded_file, original_name=''):
    batch = OneCImportBatch.objects.create(
        original_name=original_name or getattr(uploaded_file, 'name', ''),
        file=uploaded_file,
        status=OneCImportBatch.Status.PROCESSING,
    )
    parsed_rows = []
    invalid_rows = []

    try:
        if hasattr(uploaded_file, 'seek'):
            uploaded_file.seek(0)
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet = workbook.active
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            payload, error = parse_row(row, row_number)
            if payload:
                parsed_rows.append(payload)
            elif error:
                invalid_rows.append(error)
        workbook.close()

        hashes = [row['dedupe_hash'] for row in parsed_rows]
        existing_hashes = set(
            OneCEvent.objects.filter(dedupe_hash__in=hashes).values_list('dedupe_hash', flat=True)
        )
        events_to_create = [
            OneCEvent(import_batch=batch, **row)
            for row in parsed_rows
            if row['dedupe_hash'] not in existing_hashes
        ]
        OneCEvent.objects.bulk_create(events_to_create, batch_size=1000, ignore_conflicts=True)

        batch.rows_total = len(parsed_rows) + len(invalid_rows)
        batch.rows_created = len(events_to_create)
        batch.rows_skipped_duplicates = len(parsed_rows) - len(events_to_create)
        batch.rows_invalid = len(invalid_rows)
        batch.invalid_rows_preview = invalid_rows[:50]
        batch.status = OneCImportBatch.Status.DONE
        batch.finished_at = timezone.now()
        batch.save(update_fields=[
            'rows_total', 'rows_created', 'rows_skipped_duplicates', 'rows_invalid',
            'invalid_rows_preview', 'status', 'finished_at',
        ])
        logger.info(
            '1C import completed: batch=%s created=%s duplicates=%s invalid=%s',
            batch.pk, batch.rows_created, batch.rows_skipped_duplicates, batch.rows_invalid,
        )
        return batch
    except Exception as exc:
        batch.status = OneCImportBatch.Status.FAILED
        batch.error_message = str(exc)
        batch.finished_at = timezone.now()
        batch.save(update_fields=['status', 'error_message', 'finished_at'])
        logger.exception('1C import failed: batch=%s', batch.pk)
        raise


def local_date(dt):
    return timezone.localtime(dt, VLADIVOSTOK_TZ).date()


def local_hour(dt):
    return timezone.localtime(dt, VLADIVOSTOK_TZ).hour


def date_label(day):
    return day.strftime('%d.%m.%Y')


def month_label(day):
    return day.strftime('%m.%Y')


def period_caption(start_day, end_day):
    if start_day.year == end_day.year:
        if start_day.month == end_day.month:
            return f'{start_day.day}–{end_day.day} {end_day.strftime("%m.%Y")}'
        return f'{start_day.strftime("%d.%m")}–{end_day.strftime("%d.%m.%Y")}'
    return f'{start_day.strftime("%d.%m.%Y")}–{end_day.strftime("%d.%m.%Y")}'


def get_period_bounds(period_key):
    today = timezone.localdate()
    if period_key == 'last_30':
        return today - timedelta(days=29), today
    if period_key == 'current_week':
        return today - timedelta(days=today.weekday()), today
    if period_key == 'current_quarter':
        quarter_start_month = ((today.month - 1) // 3) * 3 + 1
        return date(today.year, quarter_start_month, 1), today
    if period_key == 'current_year':
        return date(today.year, 1, 1), today
    return today - timedelta(days=6), today


def iter_dates(start_day, end_day):
    days = []
    current = start_day
    while current <= end_day:
        days.append(current)
        current += timedelta(days=1)
    return days


def local_day_start(day):
    return timezone.make_aware(datetime.combine(day, time.min), VLADIVOSTOK_TZ)


def local_day_end_exclusive(day):
    return timezone.make_aware(datetime.combine(day + timedelta(days=1), time.min), VLADIVOSTOK_TZ)


def get_business_status_choices(include_other=False):
    codes = list(BUSINESS_STATUS_CODES)
    if include_other:
        codes.append(OneCEvent.StatusCode.OTHER)
    return [{'value': code, 'label': OneCEvent.StatusCode(code).label} for code in codes]


def build_anomaly_invoices(events):
    current_by_instrument = {}
    invoice_first_date = {}
    invoice_current_events = defaultdict(list)
    today = timezone.localdate()

    for event in events:
        invoice = event['invoice_number']
        event_date = local_date(event['event_datetime'])
        if invoice and invoice not in invoice_first_date:
            invoice_first_date[invoice] = event_date
        key = event['instrument_key']
        current = current_by_instrument.get(key)
        if current is None or (event['event_datetime'], event['id']) > (current['event_datetime'], current['id']):
            current_by_instrument[key] = event

    for event in current_by_instrument.values():
        if event['invoice_number']:
            invoice_current_events[event['invoice_number']].append(event)

    anomaly_invoices = set()
    anomaly_preview = []
    for invoice, current_events in invoice_current_events.items():
        first_date = invoice_first_date.get(invoice)
        if not first_date:
            continue
        age_days = (today - first_date).days
        if age_days >= 7 and all(e['status_code'] == OneCEvent.StatusCode.ACCEPTED_WAREHOUSE for e in current_events):
            anomaly_invoices.add(invoice)
            anomaly_preview.append({
                'invoice_number': invoice,
                'first_event_date': first_date.isoformat(),
                'age_days': age_days,
                'instruments_count': sum(float(e['quantity'] or 0) for e in current_events),
            })
    return anomaly_invoices, anomaly_preview, current_by_instrument


def load_all_events_for_stats():
    return list(OneCEvent.objects.order_by('event_datetime', 'id').values(
        'id', 'event_datetime', 'invoice_number', 'invoice_prefix', 'is_own_lab',
        'instrument_key', 'quantity', 'status_code', 'status_raw', 'counterparty',
        'work_description', 'barcode', 'responsible_full_name',
    ))


def is_allowed_stat_event(event, anomaly_invoices):
    return bool(event['is_own_lab'] and event['invoice_number'] not in anomaly_invoices)


def build_date_invoice_series(events, start_day, end_day, status_code, anomaly_invoices, employee=None):
    buckets = {day: set() for day in iter_dates(start_day, end_day)}
    for event in events:
        day = local_date(event['event_datetime'])
        if day < start_day or day > end_day:
            continue
        if not is_allowed_stat_event(event, anomaly_invoices):
            continue
        if event['status_code'] != status_code:
            continue
        if employee and event.get('responsible_full_name') != employee:
            continue
        buckets[day].add(event['invoice_number'])
    return [
        {
            'date': day.isoformat(),
            'label': date_label(day),
            'weekday': WEEKDAY_SHORT[day.weekday()],
            'month': month_label(day),
            'count': len(buckets[day]),
        }
        for day in buckets
    ]


def build_weekday_load(events, anomaly_invoices, period_key='current_week'):
    start_day, end_day = get_period_bounds(period_key)
    # Для витрины загрузки сравниваем только рабочие дни Пн-Пт.
    buckets = {day: set() for day in iter_dates(start_day, end_day) if day.weekday() < 5}
    tracked_statuses = {OneCEvent.StatusCode.ACCEPTED_WAREHOUSE, OneCEvent.StatusCode.ISSUED_CUSTOMER}
    for event in events:
        day = local_date(event['event_datetime'])
        if day not in buckets:
            continue
        if not is_allowed_stat_event(event, anomaly_invoices):
            continue
        if event['status_code'] not in tracked_statuses:
            continue
        counterparty = normalize_for_hash(event.get('counterparty'))
        if counterparty:
            buckets[day].add(counterparty)
    total = sum(len(values) for values in buckets.values())
    return {
        'period': period_key,
        'period_label': PERIOD_LABELS.get(period_key, PERIOD_LABELS['last_7']),
        'date_from': start_day.isoformat(),
        'date_to': end_day.isoformat(),
        'caption': f'{period_caption(start_day, end_day)} · Пн–Пт · статусы: принято/выдано',
        'total_unique_counterparty_days': total,
        'items': [
            {
                'date': day.isoformat(),
                'label': f'{WEEKDAY_SHORT[day.weekday()]} {day.strftime("%d.%m.%Y")}',
                'weekday': WEEKDAY_SHORT[day.weekday()],
                'month': month_label(day),
                'counterparties': len(buckets[day]),
                'percent': round((len(buckets[day]) / total * 100), 1) if total else 0,
            }
            for day in buckets
        ],
    }


def build_employee_chart(events, anomaly_invoices, employee='', status_code=OneCEvent.StatusCode.ACCEPTED_WAREHOUSE, period_key='last_7'):
    start_day, end_day = get_period_bounds(period_key)
    if status_code not in [choice['value'] for choice in get_business_status_choices()]:
        status_code = OneCEvent.StatusCode.ACCEPTED_WAREHOUSE
    return {
        'employee': employee,
        'status_code': status_code,
        'status_label': OneCEvent.StatusCode(status_code).label,
        'period': period_key,
        'period_label': PERIOD_LABELS.get(period_key, PERIOD_LABELS['last_7']),
        'date_from': start_day.isoformat(),
        'date_to': end_day.isoformat(),
        'caption': f'{period_caption(start_day, end_day)} · группировка по уникальным счетам',
        'items': build_date_invoice_series(events, start_day, end_day, status_code, anomaly_invoices, employee=employee or None),
    }


def build_movement_chart(events, anomaly_invoices, status_code, period_key='last_7'):
    start_day, end_day = get_period_bounds(period_key)
    return {
        'status_code': status_code,
        'status_label': OneCEvent.StatusCode(status_code).label,
        'period': period_key,
        'period_label': PERIOD_LABELS.get(period_key, PERIOD_LABELS['last_7']),
        'date_from': start_day.isoformat(),
        'date_to': end_day.isoformat(),
        'caption': f'{period_caption(start_day, end_day)} · уникальные счета по дням',
        'items': build_date_invoice_series(events, start_day, end_day, status_code, anomaly_invoices),
    }


def build_counterparty_chart(events, anomaly_invoices, counterparty='', status_mode='both', period_key='last_30'):
    start_day, end_day = get_period_bounds(period_key)
    accepted = OneCEvent.StatusCode.ACCEPTED_WAREHOUSE
    issued = OneCEvent.StatusCode.ISSUED_CUSTOMER
    if status_mode == 'accepted':
        tracked_statuses = {accepted}
        status_caption = OneCEvent.StatusCode(accepted).label
    elif status_mode == 'issued':
        tracked_statuses = {issued}
        status_caption = OneCEvent.StatusCode(issued).label
    else:
        tracked_statuses = {accepted, issued}
        status_caption = 'Принято на склад БП/СП + Выдан заказчику'

    wanted_counterparty = normalize_for_hash(counterparty)
    buckets = {day: set() for day in iter_dates(start_day, end_day)}
    for event in events:
        day = local_date(event['event_datetime'])
        if day < start_day or day > end_day:
            continue
        if not is_allowed_stat_event(event, anomaly_invoices):
            continue
        if event['status_code'] not in tracked_statuses:
            continue
        if wanted_counterparty and normalize_for_hash(event.get('counterparty')) != wanted_counterparty:
            continue
        buckets[day].add(event['instrument_key'])

    return {
        'counterparty': counterparty,
        'status_mode': status_mode,
        'status_label': status_caption,
        'period': period_key,
        'period_label': PERIOD_LABELS.get(period_key, PERIOD_LABELS['last_30']),
        'date_from': start_day.isoformat(),
        'date_to': end_day.isoformat(),
        'caption': f'{period_caption(start_day, end_day)} · уникальные приборы по дням',
        'items': [
            {
                'date': day.isoformat(),
                'label': date_label(day),
                'weekday': WEEKDAY_SHORT[day.weekday()],
                'month': month_label(day),
                'count': len(buckets[day]),
            }
            for day in buckets
        ],
    }


def get_onec_dictionaries(limit=300):
    employees = list(
        OneCEvent.objects.exclude(responsible_full_name='')
        .values('responsible_full_name')
        .annotate(events_count=Count('id'))
        .order_by('-events_count', 'responsible_full_name')[:limit]
    )
    counterparties = list(
        OneCEvent.objects.exclude(counterparty='')
        .values('counterparty')
        .annotate(events_count=Count('id'))
        .order_by('-events_count', 'counterparty')[:limit]
    )
    return {
        'employees': [
            {'value': item['responsible_full_name'], 'label': item['responsible_full_name'], 'events_count': item['events_count']}
            for item in employees
        ],
        'counterparties': [
            {'value': item['counterparty'], 'label': item['counterparty'], 'events_count': item['events_count']}
            for item in counterparties
        ],
    }


def build_onec_vitrines(params=None):
    params = params or {}
    events = load_all_events_for_stats()
    anomaly_invoices, _, _ = build_anomaly_invoices(events)

    employee = normalize_text(params.get('employee'))
    employee_status = params.get('employee_status') or OneCEvent.StatusCode.ACCEPTED_WAREHOUSE
    employee_period = params.get('employee_period') or 'last_7'
    movement_period = params.get('movement_period') or 'last_7'
    load_period = params.get('load_period') or 'current_week'
    counterparty = normalize_text(params.get('counterparty'))
    counterparty_period = params.get('counterparty_period') or 'last_30'
    counterparty_status = params.get('counterparty_status') or 'both'

    dictionaries = get_onec_dictionaries()
    return {
        'generated_at': timezone.now().astimezone(VLADIVOSTOK_TZ).isoformat(),
        'status_choices': get_business_status_choices(),
        'period_choices': [
            {'value': 'last_7', 'label': 'Последние 7 дней'},
            {'value': 'last_30', 'label': 'Последние 30 дней'},
            {'value': 'current_week', 'label': 'Текущая неделя'},
            {'value': 'current_quarter', 'label': 'Текущий квартал'},
            {'value': 'current_year', 'label': 'Текущий год'},
        ],
        'employees': dictionaries['employees'],
        'counterparties': dictionaries['counterparties'],
        'weekday_load': build_weekday_load(events, anomaly_invoices, load_period),
        'employee_chart': build_employee_chart(events, anomaly_invoices, employee, employee_status, employee_period),
        'raised_chart': build_movement_chart(events, anomaly_invoices, OneCEvent.StatusCode.TRANSFERRED_TO_LAB, movement_period),
        'lowered_chart': build_movement_chart(events, anomaly_invoices, OneCEvent.StatusCode.DONE_READY, movement_period),
        'counterparty_chart': build_counterparty_chart(events, anomaly_invoices, counterparty, counterparty_status, counterparty_period),
    }


def build_onec_stats(history_days=90):
    today = timezone.localdate()
    history_start = today - timedelta(days=history_days)
    events = load_all_events_for_stats()
    anomaly_invoices, anomaly_preview, current_by_instrument = build_anomaly_invoices(events)

    current_allowed = [event for event in current_by_instrument.values() if is_allowed_stat_event(event, anomaly_invoices)]
    today_events = [event for event in events if local_date(event['event_datetime']) == today and is_allowed_stat_event(event, anomaly_invoices)]

    today_accepted_invoices = {
        event['invoice_number'] for event in today_events if event['status_code'] == OneCEvent.StatusCode.ACCEPTED_WAREHOUSE
    }
    today_completed_invoices = {
        event['invoice_number'] for event in today_events if event['status_code'] == OneCEvent.StatusCode.DONE_READY
    }
    today_issued_invoices = {
        event['invoice_number'] for event in today_events if event['status_code'] == OneCEvent.StatusCode.ISSUED_CUSTOMER
    }

    current_status_counts = Counter()
    for event in current_allowed:
        current_status_counts[event['status_code']] += float(event['quantity'] or 0)

    hourly_invoices = defaultdict(set)
    workweek_hourly_invoices = defaultdict(set)
    for event in events:
        event_date = local_date(event['event_datetime'])
        if event_date < history_start or not is_allowed_stat_event(event, anomaly_invoices):
            continue
        if event['status_code'] == OneCEvent.StatusCode.ACCEPTED_WAREHOUSE:
            event_local = timezone.localtime(event['event_datetime'], VLADIVOSTOK_TZ)
            event_hour = event_local.hour
            invoice_number = event.get('invoice_number')
            if not invoice_number:
                continue
            hourly_invoices[event_hour].add(invoice_number)
            if event_local.weekday() < 5 and 8 <= event_hour <= 17:
                workweek_hourly_invoices[(event_local.weekday(), event_hour)].add(invoice_number)

    latest_subcontract_invoices = {
        event['invoice_number'] for event in current_by_instrument.values()
        if event['invoice_number'] and not event['is_own_lab']
    }

    return {
        'today': today.isoformat(),
        'history_days': history_days,
        'total_events': len(events),
        'total_current_instruments': len(current_by_instrument),
        'own_lab_current_instruments': len(current_allowed),
        'subcontract_current_invoices': len(latest_subcontract_invoices),
        'anomaly_invoices_count': len(anomaly_invoices),
        'anomaly_invoices_preview': sorted(anomaly_preview, key=lambda item: (-item['age_days'], item['invoice_number']))[:30],
        'today_accepted_invoices': len(today_accepted_invoices),
        'today_completed_invoices': len(today_completed_invoices),
        'today_issued_invoices': len(today_issued_invoices),
        'current_status_counts': [
            {'status_code': code, 'status_label': OneCEvent.StatusCode(code).label, 'count': count}
            for code, count in current_status_counts.most_common()
        ],
        'hourly_distribution': [
            {'hour': hour, 'count': len(hourly_invoices.get(hour, set()))}
            for hour in range(24)
        ],
        'arrival_workweek_heatmap': [
            {
                'weekday': WEEKDAY_SHORT[weekday],
                'weekday_index': weekday,
                'hours': [
                    {
                        'hour': hour,
                        'label': f'{hour:02d}:00',
                        'count': len(workweek_hourly_invoices.get((weekday, hour), set())),
                    }
                    for hour in range(8, 18)
                ],
            }
            for weekday in range(5)
        ],
    }
